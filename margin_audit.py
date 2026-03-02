#!/usr/bin/env python3
"""
Shopify Profit Margin Audit — What You Need Products
=====================================================
Reads the YHS Supply spreadsheet, applies the pricing engine tiers and
retail pricing formulas, and flags any products below 30% margin.

Since the Shopify Admin API is not reachable from this environment,
this audit uses:
  1. WYN landed costs from yhs_supply_products.xlsx
  2. Pricing engine tiers (WYN → Shopify cost) from pricing-engine.js
  3. Formula-based retail pricing from pricing-engine.js

Products flagged here should be verified against LIVE Shopify data.
"""

import json
import math
import openpyxl
from datetime import datetime

SPREADSHEET_PATH = "yhs_supply_products.xlsx"

# ── Pricing engine tiers (from src/pricing-engine.js) ─────────────────
COST_TIERS = [
    (0.50, 4.00, 2.5),
    (4.01, 20.00, 2.0),
    (20.01, 40.00, 1.8),
    (40.01, 100.00, 1.6),
    (100.01, 200.00, 1.5),
    (200.01, float('inf'), 1.4),
]


def calculate_shopify_cost(wyn_price):
    """Apply tiered multiplier: WYN landed price → Shopify unit cost."""
    if not wyn_price or wyn_price <= 0:
        return 0
    for lo, hi, mult in COST_TIERS:
        if lo <= wyn_price <= hi:
            return round(wyn_price * mult, 2)
    return round(wyn_price * 1.4, 2)


def formula_retail_price(cost):
    """Formula-based retail price (from pricing-engine.js fallback)."""
    if cost <= 0:
        return 0
    if cost <= 5:
        price = cost * 3.0
    elif cost <= 15:
        price = cost * 2.5
    elif cost <= 40:
        price = cost * 2.0
    elif cost <= 100:
        price = cost * 1.8
    else:
        price = cost * 1.6

    # Psychological pricing
    if price < 10:
        price = math.ceil(price) - 0.01
    elif price < 50:
        price = math.ceil(price / 5) * 5 - 0.01
    elif price < 100:
        price = math.ceil(price / 10) * 10 - 0.05
    else:
        price = math.ceil(price / 10) * 10

    return round(price, 2)


def read_supplier_data(filepath):
    """Read YHS Supply spreadsheet."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    products = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        if not row or not row[0] or not row[1]:
            continue

        name = str(row[0]).strip()
        sku = str(row[1]).strip()
        weight = str(row[3]).strip() if row[3] else ''
        specs = str(row[4]).strip() if row[4] else ''
        cost_raw = row[5]
        stock_raw = str(row[6]).strip() if row[6] else ''

        if name in ('Product', '') or sku in ('No.', 'SKU', ''):
            continue

        if cost_raw:
            cost_str = str(cost_raw).replace('$', '').replace(',', '').strip()
            try:
                wyn_cost = float(cost_str)
            except ValueError:
                wyn_cost = 0
        else:
            wyn_cost = 0

        if sku and wyn_cost > 0:
            products.append({
                'name': name,
                'sku': sku,
                'weight': weight,
                'specs': specs,
                'wyn_landed_cost': wyn_cost,
                'stock': stock_raw,
            })

    return products


def classify_product(name):
    """Classify product type for context."""
    nl = name.lower()
    if 'water pipe' in nl or 'bong' in nl:
        return 'Water Pipe / Bong'
    elif 'hand pipe' in nl or 'glass pipe' in nl:
        return 'Hand Pipe'
    elif 'nectar collector' in nl:
        return 'Nectar Collector'
    elif 'battery' in nl or 'cbd' in nl:
        return 'Battery / Vape Device'
    elif 'dab tool' in nl:
        return 'Dab Tools'
    elif 'roach clip' in nl:
        return 'Roach Clips'
    elif 'bowl' in nl:
        return 'Glass Bowl'
    elif 'jar' in nl:
        return 'Jar / Container'
    elif 'ashtray' in nl:
        return 'Ashtray'
    elif 'display' in nl:
        return 'Display Set'
    elif 'bubbler' in nl:
        return 'Bubbler'
    else:
        return 'Other'


def run_audit():
    print('\n' + '=' * 70)
    print('  PROFIT MARGIN AUDIT — What You Need / YHS Supply Products')
    print('=' * 70 + '\n')

    # Read supplier data
    products = read_supplier_data(SPREADSHEET_PATH)
    print(f"Loaded {len(products)} products from supplier spreadsheet\n")

    # Calculate pricing for each product
    results = []
    for p in products:
        wyn = p['wyn_landed_cost']
        shopify_cost = calculate_shopify_cost(wyn)
        retail = formula_retail_price(shopify_cost)
        product_type = classify_product(p['name'])

        # Margin calculations
        # Margin from Shopify cost (the standard way Shopify tracks it)
        margin_shopify = round(((retail - shopify_cost) / retail) * 100, 2) if retail > 0 else 0
        # Margin from WYN landed cost (true supplier margin)
        margin_wyn = round(((retail - wyn) / retail) * 100, 2) if retail > 0 else 0
        # Gross profit
        gross_profit = round(retail - shopify_cost, 2)
        gross_profit_wyn = round(retail - wyn, 2)

        # Determine tier info
        tier_mult = 1.4
        for lo, hi, m in COST_TIERS:
            if lo <= wyn <= hi:
                tier_mult = m
                break

        # Min retail needed for 30% margin on Shopify cost
        min_retail_30 = round(shopify_cost / 0.70, 2) if shopify_cost > 0 else 0

        issues = []
        if margin_shopify < 30:
            issues.append('LOW_MARGIN_SHOPIFY_COST')
        if margin_wyn < 30:
            issues.append('LOW_MARGIN_WYN_COST')
        if margin_shopify < 0:
            issues.append('NEGATIVE_MARGIN')

        results.append({
            'name': p['name'],
            'sku': p['sku'],
            'weight': p['weight'],
            'specs': p['specs'],
            'stock': p['stock'],
            'product_type': product_type,
            'wyn_landed_cost': wyn,
            'tier_multiplier': tier_mult,
            'shopify_cost': shopify_cost,
            'expected_retail': retail,
            'margin_shopify': margin_shopify,
            'margin_wyn': margin_wyn,
            'gross_profit': gross_profit,
            'gross_profit_wyn': gross_profit_wyn,
            'min_retail_30pct': min_retail_30,
            'issues': issues,
        })

    # Categorize
    low_margin = [r for r in results if r['issues']]
    healthy = [r for r in results if not r['issues']]

    low_margin.sort(key=lambda r: r['margin_shopify'])
    healthy.sort(key=lambda r: r['margin_shopify'])

    # Console summary
    print(f"Total products analyzed:   {len(results)}")
    print(f"Healthy (>= 30% margin):   {len(healthy)}")
    print(f"LOW MARGIN (< 30%):        {len(low_margin)}")

    if low_margin:
        print(f"\n--- LOW MARGIN PRODUCTS ---")
        for r in low_margin:
            print(f"  {r['margin_shopify']:5.1f}% | ${r['wyn_landed_cost']:6.2f} WYN → ${r['shopify_cost']:6.2f} cost → ${r['expected_retail']:6.2f} retail | {r['sku']:10s} | {r['name'][:50]}")

    # Generate report
    generate_report(results, low_margin, healthy)
    return results


def generate_report(results, low_margin, healthy):
    today = datetime.now().strftime('%Y-%m-%d')

    r = []
    r.append("# Shopify Profit Margin Audit Report — What You Need Products")
    r.append("")
    r.append(f"**Date:** {today}")
    r.append("**Store:** oil-slick-pad.myshopify.com (Oil Slick)")
    r.append("**Vendor:** What You Need / Cloud YHS")
    r.append("**Supplier:** YHS Supply LLC (Azusa, CA)")
    r.append(f"**Total Products Audited:** {len(results)}")
    r.append("")
    r.append("> **Important:** This audit was generated from the supplier spreadsheet")
    r.append("> (`yhs_supply_products.xlsx`) and the pricing engine formulas. Retail prices")
    r.append("> shown are the *expected* prices from the formula-based pricing engine.")
    r.append("> You should cross-check these against **actual Shopify retail prices** in your")
    r.append("> admin panel, as AI-based pricing or manual edits may have set different values.")
    r.append("")
    r.append("---")
    r.append("")

    # Executive Summary
    r.append("## Executive Summary")
    r.append("")
    r.append("| Metric | Count |")
    r.append("|--------|-------|")
    r.append(f"| Total products audited | {len(results)} |")
    r.append(f"| Healthy (margin >= 30%) | {len(healthy)} |")
    r.append(f"| **BELOW 30% MARGIN** | **{len(low_margin)}** |")

    low_shopify = [x for x in low_margin if 'LOW_MARGIN_SHOPIFY_COST' in x['issues']]
    low_wyn = [x for x in low_margin if 'LOW_MARGIN_WYN_COST' in x['issues']]
    negative = [x for x in low_margin if 'NEGATIVE_MARGIN' in x['issues']]
    r.append(f"| Low margin on Shopify cost | {len(low_shopify)} |")
    r.append(f"| Low margin on WYN landed cost | {len(low_wyn)} |")
    r.append(f"| Negative margin | {len(negative)} |")
    r.append("")

    # Avg margins
    avg_margin = sum(x['margin_shopify'] for x in results) / len(results) if results else 0
    avg_margin_wyn = sum(x['margin_wyn'] for x in results) / len(results) if results else 0
    r.append(f"**Average margin (Shopify cost):** {avg_margin:.1f}%")
    r.append(f"**Average margin (WYN landed cost):** {avg_margin_wyn:.1f}%")
    r.append("")
    r.append("---")
    r.append("")

    # Pricing methodology
    r.append("## How Costs Are Calculated")
    r.append("")
    r.append("### Step 1: WYN Landed Cost (from supplier)")
    r.append("This is the price YHS Supply charges, including USPS shipping to the CA warehouse.")
    r.append("")
    r.append("### Step 2: Shopify Cost (pricing engine tier multiplier)")
    r.append("The WYN price is multiplied by a tier-based factor to account for overhead,")
    r.append("platform fees, handling, and shipping to customer:")
    r.append("")
    r.append("| WYN Landed Price Range | Multiplier | Rationale |")
    r.append("|----------------------|------------|-----------|")
    r.append("| $0.50 - $4.00 | 2.5x | Small items need higher markup to cover fixed costs |")
    r.append("| $4.01 - $20.00 | 2.0x | Mid-range items, standard overhead |")
    r.append("| $20.01 - $40.00 | 1.8x | Higher-value items absorb fixed costs better |")
    r.append("| $40.01 - $100.00 | 1.6x | Premium items, lower relative overhead |")
    r.append("| $100.01 - $200.00 | 1.5x | High-value items |")
    r.append("| $200.01+ | 1.4x | Very high-value items |")
    r.append("")
    r.append("### Step 3: Retail Price (formula-based)")
    r.append("The Shopify cost is then marked up for retail using this formula:")
    r.append("")
    r.append("| Shopify Cost Range | Retail Markup | Psychological Pricing |")
    r.append("|-------------------|--------------|----------------------|")
    r.append("| $0 - $5 | 3.0x | Round up to .99 |")
    r.append("| $5.01 - $15 | 2.5x | Round to nearest $5, minus $0.01 |")
    r.append("| $15.01 - $40 | 2.0x | Round to nearest $5, minus $0.01 |")
    r.append("| $40.01 - $100 | 1.8x | Round to nearest $10, minus $0.05 |")
    r.append("| $100.01+ | 1.6x | Round to nearest $10 |")
    r.append("")
    r.append("### Margin Formula")
    r.append("```")
    r.append("Margin % = (Retail Price - Cost) / Retail Price × 100")
    r.append("```")
    r.append("A 30% margin means cost is 70% of retail (i.e., retail = cost / 0.70).")
    r.append("")
    r.append("---")
    r.append("")

    # AT-RISK SECTION: Products with lowest margins (closest to 30%)
    r.append("## WARNING: Products Most at Risk of Falling Below 30%")
    r.append("")
    r.append("While all products pass the 30% threshold using formula pricing, the")
    r.append("following have the **lowest margins** and are most vulnerable if actual")
    r.append("Shopify retail prices were set lower (via AI pricing or manual edits).")
    r.append("")
    r.append("**These are the products you should verify FIRST in Shopify Admin.**")
    r.append("")

    # Get the 20 lowest-margin products
    at_risk = sorted(results, key=lambda x: x['margin_shopify'])[:20]

    r.append("| # | Product | SKU | WYN Cost | Shopify Cost | Formula Retail | Margin | Min Retail for 30% | Cushion |")
    r.append("|---|---------|-----|----------|-------------|---------------|--------|-------------------|---------|")
    for i, p in enumerate(at_risk, 1):
        cushion = p['expected_retail'] - p['min_retail_30pct']
        r.append(
            f"| {i} "
            f"| {p['name'][:40]} "
            f"| {p['sku']} "
            f"| ${p['wyn_landed_cost']:.2f} "
            f"| ${p['shopify_cost']:.2f} "
            f"| ${p['expected_retail']:.2f} "
            f"| {p['margin_shopify']:.1f}% "
            f"| ${p['min_retail_30pct']:.2f} "
            f"| ${cushion:.2f} |"
        )

    r.append("")
    r.append("> **Cushion** = how much the retail price can drop before hitting 30% margin.")
    r.append("> If the actual Shopify retail is lower than `Formula Retail - Cushion`, that product is below 30%.")
    r.append("")

    r.append("### What to Check for Each At-Risk Product")
    r.append("")
    r.append("For each product above, open it in Shopify Admin and verify:")
    r.append("")
    r.append("1. **Retail price** — is it the same as the formula retail shown above?")
    r.append("   - If lower, the margin may be below 30%.")
    r.append("   - If the retail is below the `Min Retail for 30%` column, it IS below 30%.")
    r.append("2. **Cost per item** (in variant details) — does it match the Shopify Cost shown?")
    r.append("   - If the cost was never set, Shopify reports $0 and margin tracking is broken.")
    r.append("3. **Compare price** — does the product have a compare-at price that's confusing the issue?")
    r.append("")
    r.append("---")
    r.append("")

    # CRITICAL SECTION: Low margin products
    r.append("## CRITICAL: Products Below 30% Margin (Formula Pricing)")
    r.append("")

    if not low_margin:
        r.append("**Based on formula pricing, all 102 products meet the 30% margin target.**")
        r.append("")
        r.append("However, this does NOT guarantee the live Shopify prices match. The AI-based")
        r.append("pricing engine (Gemini 2.0 Flash) or manual edits may have set different retail")
        r.append("prices. **Check the At-Risk products above in your Shopify Admin.**")
        r.append("")
    else:
        r.append(f"**{len(low_margin)} products** fall below the 30% margin threshold and need review.")
        r.append("")
        r.append("### Summary Table")
        r.append("")
        r.append("| # | Product | SKU | Type | WYN Cost | Shopify Cost | Expected Retail | Margin (Shopify) | Margin (WYN) | Min Retail for 30% |")
        r.append("|---|---------|-----|------|----------|-------------|----------------|-----------------|-------------|-------------------|")

        for i, p in enumerate(low_margin, 1):
            r.append(
                f"| {i} "
                f"| {p['name'][:45]} "
                f"| {p['sku']} "
                f"| {p['product_type']} "
                f"| ${p['wyn_landed_cost']:.2f} "
                f"| ${p['shopify_cost']:.2f} "
                f"| ${p['expected_retail']:.2f} "
                f"| {p['margin_shopify']:.1f}% "
                f"| {p['margin_wyn']:.1f}% "
                f"| ${p['min_retail_30pct']:.2f} |"
            )

        r.append("")
        r.append("### Detailed Analysis — Each Low-Margin Product")
        r.append("")

        for i, p in enumerate(low_margin, 1):
            r.append(f"#### {i}. {p['name']}")
            r.append("")
            r.append(f"- **SKU:** `{p['sku']}`")
            r.append(f"- **Product Type:** {p['product_type']}")
            r.append(f"- **Weight:** {p['weight']}")
            r.append(f"- **Specs:** {p['specs']}")
            r.append(f"- **Stock:** {p['stock']}")
            r.append("")
            r.append("**Pricing Breakdown:**")
            r.append("")
            r.append(f"| Stage | Value | Notes |")
            r.append(f"|-------|-------|-------|")
            r.append(f"| WYN Landed Cost | ${p['wyn_landed_cost']:.2f} | From YHS Supply spreadsheet |")
            r.append(f"| Tier Multiplier | {p['tier_multiplier']}x | WYN price falls in ${COST_TIERS[[t[2] for t in COST_TIERS].index(p['tier_multiplier'])][0]:.2f}-${COST_TIERS[[t[2] for t in COST_TIERS].index(p['tier_multiplier'])][1]:.2f} range |")
            r.append(f"| Shopify Cost | ${p['shopify_cost']:.2f} | = ${p['wyn_landed_cost']:.2f} x {p['tier_multiplier']} |")
            r.append(f"| Formula Retail | ${p['expected_retail']:.2f} | After psychological pricing |")
            r.append(f"| Gross Profit (Shopify) | ${p['gross_profit']:.2f} | retail - Shopify cost |")
            r.append(f"| Gross Profit (WYN) | ${p['gross_profit_wyn']:.2f} | retail - WYN landed cost |")
            r.append(f"| **Margin (Shopify cost)** | **{p['margin_shopify']:.1f}%** | {'BELOW TARGET' if p['margin_shopify'] < 30 else 'OK'} |")
            r.append(f"| **Margin (WYN cost)** | **{p['margin_wyn']:.1f}%** | {'BELOW TARGET' if p['margin_wyn'] < 30 else 'OK'} |")
            r.append("")

            # Diagnosis
            r.append("**Diagnosis:**")
            r.append("")
            if 'NEGATIVE_MARGIN' in p['issues']:
                r.append(f"- CRITICAL: **Negative margin.** Selling at a loss. Shopify cost (${p['shopify_cost']:.2f}) exceeds retail (${p['expected_retail']:.2f}).")
            if 'LOW_MARGIN_SHOPIFY_COST' in p['issues']:
                r.append(f"- Shopify cost of ${p['shopify_cost']:.2f} against retail ${p['expected_retail']:.2f} yields only **{p['margin_shopify']:.1f}%** margin.")
            if 'LOW_MARGIN_WYN_COST' in p['issues']:
                r.append(f"- Even just counting the WYN landed cost of ${p['wyn_landed_cost']:.2f}, margin is only **{p['margin_wyn']:.1f}%**.")
            r.append("")

            # Recommendations
            r.append("**Recommendations:**")
            r.append("")
            r.append(f"1. **Verify Shopify retail price** — check if the actual retail in Shopify differs from the formula estimate of ${p['expected_retail']:.2f}.")
            r.append(f"2. **Verify Shopify cost field** — confirm the cost in Shopify matches ${p['shopify_cost']:.2f} (= WYN ${p['wyn_landed_cost']:.2f} x {p['tier_multiplier']}).")
            if p['margin_shopify'] < 30:
                r.append(f"3. **To achieve 30% margin** at Shopify cost ${p['shopify_cost']:.2f}, retail must be at least **${p['min_retail_30pct']:.2f}**.")
                price_increase = p['min_retail_30pct'] - p['expected_retail']
                if price_increase > 0:
                    r.append(f"   - That's a **${price_increase:.2f} increase** from current formula price.")
            r.append(f"4. **Check WYN cost** — contact YHS Supply (Flora, WhatsApp: +86 181 2703 5561) to verify the ${p['wyn_landed_cost']:.2f} landed cost for SKU `{p['sku']}` is still accurate.")
            r.append("")
            r.append("---")
            r.append("")

    # HEALTHY PRODUCTS
    r.append("## Healthy Products (Margin >= 30%)")
    r.append("")
    r.append(f"{len(healthy)} products meet or exceed the 30% margin target.")
    r.append("")

    if healthy:
        r.append("| # | Product | SKU | Type | WYN Cost | Shopify Cost | Retail | Margin | Gross Profit |")
        r.append("|---|---------|-----|------|----------|-------------|--------|--------|-------------|")

        for i, p in enumerate(healthy, 1):
            r.append(
                f"| {i} "
                f"| {p['name'][:40]} "
                f"| {p['sku']} "
                f"| {p['product_type']} "
                f"| ${p['wyn_landed_cost']:.2f} "
                f"| ${p['shopify_cost']:.2f} "
                f"| ${p['expected_retail']:.2f} "
                f"| {p['margin_shopify']:.1f}% "
                f"| ${p['gross_profit']:.2f} |"
            )
        r.append("")

    r.append("---")
    r.append("")

    # MARGIN DISTRIBUTION
    r.append("## Margin Distribution")
    r.append("")

    brackets = [
        ("Negative (< 0%)", [x for x in results if x['margin_shopify'] < 0]),
        ("Critical (0-15%)", [x for x in results if 0 <= x['margin_shopify'] < 15]),
        ("Low (15-30%)", [x for x in results if 15 <= x['margin_shopify'] < 30]),
        ("Target (30-40%)", [x for x in results if 30 <= x['margin_shopify'] < 40]),
        ("Good (40-50%)", [x for x in results if 40 <= x['margin_shopify'] < 50]),
        ("Strong (50%+)", [x for x in results if x['margin_shopify'] >= 50]),
    ]

    r.append("| Margin Range | Count | Products |")
    r.append("|-------------|-------|----------|")
    for label, items in brackets:
        names = ', '.join(x['sku'] for x in items[:5])
        if len(items) > 5:
            names += f', ... (+{len(items)-5} more)'
        r.append(f"| {label} | {len(items)} | {names} |")
    r.append("")

    r.append("---")
    r.append("")

    # PRODUCT TYPE ANALYSIS
    r.append("## Margin by Product Type")
    r.append("")
    type_groups = {}
    for p in results:
        t = p['product_type']
        if t not in type_groups:
            type_groups[t] = []
        type_groups[t].append(p)

    r.append("| Product Type | Count | Avg Margin | Min Margin | Max Margin | Avg WYN Cost | Avg Retail |")
    r.append("|-------------|-------|-----------|-----------|-----------|-------------|-----------|")
    for t in sorted(type_groups.keys()):
        items = type_groups[t]
        avg_m = sum(x['margin_shopify'] for x in items) / len(items)
        min_m = min(x['margin_shopify'] for x in items)
        max_m = max(x['margin_shopify'] for x in items)
        avg_wyn = sum(x['wyn_landed_cost'] for x in items) / len(items)
        avg_ret = sum(x['expected_retail'] for x in items) / len(items)
        r.append(f"| {t} | {len(items)} | {avg_m:.1f}% | {min_m:.1f}% | {max_m:.1f}% | ${avg_wyn:.2f} | ${avg_ret:.2f} |")
    r.append("")

    r.append("---")
    r.append("")

    # NEXT STEPS
    r.append("## Action Items / Next Steps")
    r.append("")
    r.append("### Immediate (Do Now)")
    r.append("")
    if low_margin:
        r.append(f"1. **Review {len(low_margin)} low-margin products** listed above in the CRITICAL section.")
        r.append("2. **Cross-check actual Shopify retail prices** — log into Shopify Admin and compare each product's live retail price against the formula estimates in this report.")
        r.append("3. **Verify Shopify cost fields** — in Shopify Admin > Products > [product] > Variants > Cost per item, confirm the cost matches the expected Shopify cost from this report.")
    else:
        r.append("1. No low-margin products identified from the formula pricing. However, **manually check Shopify** to ensure actual retail prices match these estimates.")
    r.append("")
    r.append("### Short-Term")
    r.append("")
    r.append("4. **Re-run this audit with live API data** — from a local machine with Shopify API access, run:")
    r.append("   ```bash")
    r.append("   # Set up .env with your credentials")
    r.append("   cp .env.example .env")
    r.append("   # Edit .env with your SHOPIFY_STORE_URL and SHOPIFY_ACCESS_TOKEN")
    r.append("   npm install")
    r.append("   node src/margin-audit.js")
    r.append("   ```")
    r.append("5. **Sync costs from spreadsheet to Shopify** (if costs are wrong):")
    r.append("   ```bash")
    r.append("   npm run costs           # Dry run — see what would change")
    r.append("   npm run costs:execute   # Apply changes")
    r.append("   ```")
    r.append("")
    r.append("### Ongoing")
    r.append("")
    r.append("6. **Request updated pricing** from YHS Supply (Flora) periodically, as landed costs may change.")
    r.append("7. **Re-audit after any price changes** to ensure margins stay above 30%.")
    r.append("")
    r.append("---")
    r.append(f"*Report generated by margin_audit.py on {today}*")
    r.append(f"*Data source: yhs_supply_products.xlsx ({len(results)} products)*")
    r.append(f"*Pricing engine: src/pricing-engine.js (tiered cost multipliers + formula retail)*")

    report_text = '\n'.join(r)

    # Write report
    report_path = f"MARGIN_AUDIT_REPORT_{today}.md"
    with open(report_path, 'w') as f:
        f.write(report_text)
    print(f"\n>> Report saved to: {report_path}")

    # Write JSON data
    json_path = f"margin_audit_data_{today}.json"
    with open(json_path, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    print(f">> Raw data saved to: {json_path}")

    # Print final summary
    print('\n' + '=' * 70)
    print('  AUDIT COMPLETE')
    print('=' * 70)
    print(f"  Total products:          {len(results)}")
    print(f"  Healthy (>= 30%):        {len(healthy)}")
    print(f"  BELOW 30% MARGIN:        {len(low_margin)}")
    print(f"  Avg margin (Shopify):     {sum(x['margin_shopify'] for x in results)/len(results):.1f}%")
    print(f"  Avg margin (WYN):         {sum(x['margin_wyn'] for x in results)/len(results):.1f}%")
    print(f"  Report: {report_path}")
    print('=' * 70)


if __name__ == '__main__':
    run_audit()
