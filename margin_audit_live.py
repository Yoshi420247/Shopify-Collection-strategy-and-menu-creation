#!/usr/bin/env python3
"""
Full Shopify + WooCommerce Cross-Reference Audit
=================================================
1. Fetches ALL 'What You Need' vendor products from Shopify (all pages)
2. Fetches inventory item costs for each variant
3. Reads supplier spreadsheet for WYN landed costs
4. Cross-references everything and calculates margins
5. Flags products below 30% margin
"""

import json
import os
import re
import time
import urllib.request
import urllib.error
import ssl
import openpyxl
from datetime import datetime

# ── Config (reads from .env or environment variables) ─────────────────
STORE_URL = os.environ.get("SHOPIFY_STORE_URL", "oil-slick-pad.myshopify.com")
ACCESS_TOKEN = os.environ.get("SHOPIFY_ACCESS_TOKEN", "")
API_VERSION = os.environ.get("SHOPIFY_API_VERSION", "2024-01")
BASE_URL = f"https://{STORE_URL}/admin/api/{API_VERSION}"

if not ACCESS_TOKEN:
    # Try reading from .env file
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith('SHOPIFY_ACCESS_TOKEN='):
                    ACCESS_TOKEN = line.split('=', 1)[1].strip()
    if not ACCESS_TOKEN:
        print("ERROR: SHOPIFY_ACCESS_TOKEN not set. Set it in .env or as an environment variable.")
        exit(1)

# Pricing engine tiers
COST_TIERS = [
    (0.50, 4.00, 2.5),
    (4.01, 20.00, 2.0),
    (20.01, 40.00, 1.8),
    (40.01, 100.00, 1.6),
    (100.01, 200.00, 1.5),
    (200.01, float('inf'), 1.4),
]


def calc_expected_cost(wyn_price):
    if not wyn_price or wyn_price <= 0:
        return 0
    for lo, hi, mult in COST_TIERS:
        if lo <= wyn_price <= hi:
            return round(wyn_price * mult, 2)
    return round(wyn_price * 1.4, 2)


def shopify_get(url, retries=4):
    """Shopify API GET with retries."""
    if not url.startswith('http'):
        url = f"{BASE_URL}/{url}"

    headers = {
        'X-Shopify-Access-Token': ACCESS_TOKEN,
        'Content-Type': 'application/json',
    }
    ctx = ssl.create_default_context()

    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(url, headers=headers)
            resp = urllib.request.urlopen(req, timeout=30, context=ctx)
            data = json.loads(resp.read().decode())
            link_header = resp.headers.get('Link', '')
            return data, link_header
        except Exception as e:
            if attempt < retries:
                wait = 2 ** attempt
                print(f"    Retry {attempt}/{retries} in {wait}s... ({e})")
                time.sleep(wait)
            else:
                raise


def fetch_all_wyn_products():
    """Fetch all What You Need vendor products with cursor pagination."""
    all_products = []
    url = "products.json?limit=250&vendor=What+You+Need"
    page = 0

    while url:
        page += 1
        data, link_header = shopify_get(url)
        products = data.get('products', [])
        if not products:
            break
        all_products.extend(products)
        print(f"  Page {page}: {len(products)} products (total: {len(all_products)})")

        # Parse next page URL from Link header
        url = None
        if link_header:
            match = re.search(r'<([^>]+)>;\s*rel="next"', link_header)
            if match:
                url = match.group(1)
        time.sleep(0.5)

    return all_products


def fetch_inventory_costs_batch(inventory_item_ids):
    """Fetch costs for up to 100 inventory items at once."""
    costs = {}
    # Shopify allows up to 100 IDs per batch
    for i in range(0, len(inventory_item_ids), 100):
        batch = inventory_item_ids[i:i+100]
        ids_str = ','.join(str(x) for x in batch)
        url = f"inventory_items.json?ids={ids_str}"

        try:
            data, _ = shopify_get(url)
            for item in data.get('inventory_items', []):
                cost = item.get('cost')
                costs[item['id']] = float(cost) if cost else None
        except Exception as e:
            print(f"  Warning: batch fetch failed: {e}")

        time.sleep(0.5)

    return costs


def read_supplier_costs(filepath):
    """Read YHS Supply spreadsheet."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    products = {}

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        name = str(row[0]).strip()
        sku = str(row[1]).strip()
        cost_raw = row[5]

        if name in ('Product', '') or sku in ('No.', 'SKU', ''):
            continue

        if cost_raw:
            try:
                cost = float(str(cost_raw).replace('$', '').replace(',', '').strip())
            except ValueError:
                cost = 0
        else:
            cost = 0

        if sku and cost > 0:
            products[sku.upper()] = {
                'name': name,
                'sku': sku,
                'wyn_landed_cost': cost,
            }

    return products


def main():
    print('\n' + '=' * 70)
    print('  FULL MARGIN AUDIT — Shopify Live Data + Supplier Cross-Reference')
    print('=' * 70 + '\n')

    # Step 1: Supplier spreadsheet
    print("Step 1: Reading supplier spreadsheet...")
    supplier = read_supplier_costs('yhs_supply_products.xlsx')
    print(f"  Loaded {len(supplier)} supplier products\n")

    # Step 2: Fetch all WYN products from Shopify
    print("Step 2: Fetching all 'What You Need' products from Shopify...")
    all_products = fetch_all_wyn_products()
    print(f"  Total products: {len(all_products)}\n")

    # Step 3: Collect all inventory item IDs
    print("Step 3: Collecting inventory item IDs...")
    all_variants = []
    inv_ids = []
    for p in all_products:
        for v in p.get('variants', []):
            inv_id = v.get('inventory_item_id')
            if inv_id:
                inv_ids.append(inv_id)
            all_variants.append({
                'product_title': p['title'],
                'variant_title': v.get('title', '') if v.get('title') != 'Default Title' else '',
                'sku': (v.get('sku') or '').strip(),
                'price': float(v.get('price') or 0),
                'compare_at_price': float(v['compare_at_price']) if v.get('compare_at_price') else None,
                'inventory_item_id': inv_id,
                'product_id': p['id'],
                'variant_id': v['id'],
                'vendor': p.get('vendor', ''),
                'tags': p.get('tags', ''),
            })
    print(f"  Total variants: {len(all_variants)}, inventory items: {len(inv_ids)}\n")

    # Step 4: Batch fetch inventory costs
    print("Step 4: Fetching inventory costs (this may take a minute)...")
    costs = fetch_inventory_costs_batch(inv_ids)
    print(f"  Retrieved costs for {sum(1 for c in costs.values() if c is not None)} items\n")

    # Step 5: Calculate margins
    print("Step 5: Calculating margins...")
    results = []
    for v in all_variants:
        inv_id = v['inventory_item_id']
        shopify_cost = costs.get(inv_id)
        sku = v['sku']
        retail = v['price']
        compare_at = v['compare_at_price']

        # Supplier lookup
        sup = supplier.get(sku.upper()) if sku else None
        wyn_cost = sup['wyn_landed_cost'] if sup else None
        expected_cost = calc_expected_cost(wyn_cost) if wyn_cost else None

        # Margins
        margin_shopify = None
        if shopify_cost and shopify_cost > 0 and retail > 0:
            margin_shopify = round(((retail - shopify_cost) / retail) * 100, 2)

        margin_wyn = None
        if wyn_cost and retail > 0:
            margin_wyn = round(((retail - wyn_cost) / retail) * 100, 2)

        # Issues
        issues = []
        if retail == 0:
            issues.append('ZERO_PRICE')
        if shopify_cost is None or shopify_cost == 0:
            issues.append('NO_COST')
        if margin_shopify is not None and margin_shopify < 30:
            issues.append('LOW_MARGIN')
        if margin_shopify is not None and margin_shopify < 0:
            issues.append('NEGATIVE_MARGIN')
        if shopify_cost and expected_cost and abs(shopify_cost - expected_cost) > 1:
            issues.append('COST_MISMATCH')
        if not sup and sku and not sku.startswith('NOSKU'):
            issues.append('NO_SUPPLIER')

        min_retail_30 = round(shopify_cost / 0.7, 2) if shopify_cost and shopify_cost > 0 else None

        results.append({
            'product_title': v['product_title'],
            'variant_title': v['variant_title'],
            'sku': sku or 'N/A',
            'retail_price': retail,
            'compare_at_price': compare_at,
            'shopify_cost': shopify_cost,
            'wyn_landed_cost': wyn_cost,
            'expected_shopify_cost': expected_cost,
            'margin_shopify': margin_shopify,
            'margin_wyn': margin_wyn,
            'min_retail_30': min_retail_30,
            'product_id': v['product_id'],
            'variant_id': v['variant_id'],
            'inventory_item_id': inv_id,
            'vendor': v['vendor'],
            'tags': v['tags'],
            'issues': issues,
        })

    print(f"  Processed {len(results)} variants\n")

    # Step 6: Categorize and generate report
    low_margin = [r for r in results if 'LOW_MARGIN' in r['issues']]
    negative = [r for r in results if 'NEGATIVE_MARGIN' in r['issues']]
    no_cost = [r for r in results if 'NO_COST' in r['issues']]
    zero_price = [r for r in results if 'ZERO_PRICE' in r['issues']]
    cost_mismatch = [r for r in results if 'COST_MISMATCH' in r['issues']]
    no_supplier = [r for r in results if 'NO_SUPPLIER' in r['issues']]
    healthy = [r for r in results if not r['issues']]

    # With cost set and calculable margin
    has_margin = [r for r in results if r['margin_shopify'] is not None]

    print('=' * 70)
    print('  AUDIT SUMMARY')
    print('=' * 70)
    print(f"  Total variants:           {len(results)}")
    print(f"  With calculable margin:   {len(has_margin)}")
    print(f"  Healthy (>= 30%):         {len(healthy)}")
    print(f"  LOW MARGIN (< 30%):       {len(low_margin)}")
    print(f"  NEGATIVE MARGIN:          {len(negative)}")
    print(f"  Zero retail price:        {len(zero_price)}")
    print(f"  No cost set in Shopify:   {len(no_cost)}")
    print(f"  Cost mismatch:            {len(cost_mismatch)}")
    print(f"  No supplier match:        {len(no_supplier)}")

    if has_margin:
        avg_m = sum(r['margin_shopify'] for r in has_margin) / len(has_margin)
        print(f"  Avg margin (with cost):   {avg_m:.1f}%")
    print('=' * 70 + '\n')

    # Print low-margin products
    if low_margin:
        print(f"\n--- {len(low_margin)} LOW MARGIN PRODUCTS (< 30%) ---")
        low_margin.sort(key=lambda r: r['margin_shopify'] if r['margin_shopify'] is not None else 999)
        for r in low_margin:
            name = r['product_title'][:45]
            if r['variant_title']:
                name += f" ({r['variant_title'][:15]})"
            print(f"  {r['margin_shopify']:6.1f}% | ${r['shopify_cost']:7.2f} cost | ${r['retail_price']:7.2f} retail | {r['sku']:>12} | {name}")

    if negative:
        print(f"\n--- {len(negative)} NEGATIVE MARGIN PRODUCTS ---")
        for r in negative:
            name = r['product_title'][:45]
            print(f"  {r['margin_shopify']:6.1f}% | ${r['shopify_cost']:7.2f} cost | ${r['retail_price']:7.2f} retail | {r['sku']:>12} | {name}")

    # Generate report
    generate_report(results, low_margin, negative, no_cost, zero_price, cost_mismatch, no_supplier, healthy, has_margin, supplier)

    return results


def generate_report(results, low_margin, negative, no_cost, zero_price, cost_mismatch, no_supplier, healthy, has_margin, supplier):
    today = datetime.now().strftime('%Y-%m-%d')
    r = []

    r.append("# Shopify Profit Margin Audit Report — LIVE DATA")
    r.append("")
    r.append(f"**Date:** {today}")
    r.append("**Store:** oil-slick-pad.myshopify.com (Oil Slick)")
    r.append("**Vendor Filter:** What You Need")
    r.append(f"**Total Products:** {len(set(x['product_id'] for x in results))}")
    r.append(f"**Total Variants:** {len(results)}")
    r.append("")
    r.append("> This audit uses **live Shopify data** — actual retail prices and cost fields")
    r.append("> from the Shopify Admin API, cross-referenced with the YHS Supply spreadsheet.")
    r.append("")
    r.append("---")
    r.append("")

    # Executive Summary
    r.append("## Executive Summary")
    r.append("")
    r.append("| Metric | Count |")
    r.append("|--------|-------|")
    r.append(f"| Total variants audited | {len(results)} |")
    r.append(f"| Variants with calculable margin | {len(has_margin)} |")
    r.append(f"| Healthy (margin >= 30%) | {len(healthy)} |")
    r.append(f"| **LOW MARGIN (< 30%)** | **{len(low_margin)}** |")
    r.append(f"| **NEGATIVE MARGIN (losing money)** | **{len(negative)}** |")
    r.append(f"| Zero retail price ($0.00) | {len(zero_price)} |")
    r.append(f"| Missing Shopify cost field | {len(no_cost)} |")
    r.append(f"| Cost mismatch vs supplier | {len(cost_mismatch)} |")
    r.append(f"| No supplier spreadsheet match | {len(no_supplier)} |")
    r.append("")

    if has_margin:
        avg_m = sum(x['margin_shopify'] for x in has_margin) / len(has_margin)
        min_m = min(x['margin_shopify'] for x in has_margin)
        max_m = max(x['margin_shopify'] for x in has_margin)
        r.append(f"**Average margin:** {avg_m:.1f}% | **Min:** {min_m:.1f}% | **Max:** {max_m:.1f}%")
    r.append("")
    r.append("---")
    r.append("")

    # CRITICAL: Low margin
    r.append("## CRITICAL: Products Below 30% Margin")
    r.append("")

    if not low_margin:
        r.append("**No products are below 30% margin.** All pricing is healthy.")
        r.append("")
    else:
        r.append(f"**{len(low_margin)} variants** are below the 30% margin target and need review.")
        r.append("")

        low_margin.sort(key=lambda x: x['margin_shopify'] if x['margin_shopify'] is not None else 999)

        r.append("| # | Product | SKU | Retail | Shopify Cost | Margin | WYN Cost | Expected Cost | Min Retail @30% |")
        r.append("|---|---------|-----|--------|-------------|--------|----------|---------------|----------------|")
        for i, p in enumerate(low_margin, 1):
            name = p['product_title'][:40]
            if p['variant_title']:
                name += f" ({p['variant_title'][:15]})"
            cost_str = f"${p['shopify_cost']:.2f}" if p['shopify_cost'] is not None else "NOT SET"
            wyn_str = f"${p['wyn_landed_cost']:.2f}" if p['wyn_landed_cost'] else "N/A"
            exp_str = f"${p['expected_shopify_cost']:.2f}" if p['expected_shopify_cost'] else "N/A"
            min_str = f"${p['min_retail_30']:.2f}" if p['min_retail_30'] else "N/A"
            margin_str = f"{p['margin_shopify']:.1f}%" if p['margin_shopify'] is not None else "N/A"
            r.append(f"| {i} | {name} | {p['sku']} | ${p['retail_price']:.2f} | {cost_str} | {margin_str} | {wyn_str} | {exp_str} | {min_str} |")
        r.append("")

        # Detailed breakdown
        r.append("### Detailed Analysis — Each Low-Margin Product")
        r.append("")
        for p in low_margin:
            name = p['product_title']
            if p['variant_title']:
                name += f" ({p['variant_title']})"
            r.append(f"#### {name}")
            r.append("")
            r.append(f"- **SKU:** `{p['sku']}`")
            r.append(f"- **Product ID:** {p['product_id']} | **Variant ID:** {p['variant_id']}")
            r.append(f"- **Retail Price:** ${p['retail_price']:.2f}")
            if p['compare_at_price']:
                r.append(f"- **Compare At Price:** ${p['compare_at_price']:.2f}")
            if p['shopify_cost'] is not None:
                r.append(f"- **Shopify Cost:** ${p['shopify_cost']:.2f}")
            else:
                r.append(f"- **Shopify Cost:** NOT SET")
            if p['wyn_landed_cost']:
                r.append(f"- **WYN Landed Cost:** ${p['wyn_landed_cost']:.2f}")
            if p['expected_shopify_cost']:
                r.append(f"- **Expected Shopify Cost (pricing engine):** ${p['expected_shopify_cost']:.2f}")
            if p['margin_shopify'] is not None:
                r.append(f"- **Margin:** {p['margin_shopify']:.1f}%")

            # Recommendation
            if p['shopify_cost'] and p['shopify_cost'] > 0:
                needed = round(p['shopify_cost'] / 0.7, 2)
                r.append(f"- **To achieve 30% margin:** Raise retail to at least **${needed:.2f}** (currently ${p['retail_price']:.2f})")
                diff = needed - p['retail_price']
                if diff > 0:
                    r.append(f"  - Required increase: **+${diff:.2f}**")

            # Check cost correctness
            if p['wyn_landed_cost'] and p['shopify_cost']:
                exp = calc_expected_cost(p['wyn_landed_cost'])
                if abs(p['shopify_cost'] - exp) > 1:
                    r.append(f"- **Cost Alert:** Shopify cost (${p['shopify_cost']:.2f}) doesn't match expected (${exp:.2f}) from WYN ${p['wyn_landed_cost']:.2f}")

            r.append("")

    r.append("---")
    r.append("")

    # NEGATIVE MARGIN
    if negative:
        r.append("## URGENT: Negative Margin Products (Selling at a Loss)")
        r.append("")
        r.append(f"**{len(negative)} variants** are priced below their cost.")
        r.append("")
        for p in negative:
            name = p['product_title']
            r.append(f"- **{name}** | SKU: {p['sku']} | Retail: ${p['retail_price']:.2f} | Cost: ${p['shopify_cost']:.2f} | Margin: {p['margin_shopify']:.1f}%")
        r.append("")
        r.append("---")
        r.append("")

    # ZERO PRICE
    if zero_price:
        r.append("## Products with $0.00 Retail Price")
        r.append("")
        r.append(f"**{len(zero_price)} variants** have a $0.00 retail price. These may be draft/unpublished or incorrectly configured.")
        r.append("")
        r.append("| Product | SKU | Cost | Tags |")
        r.append("|---------|-----|------|------|")
        for p in zero_price[:50]:  # Cap at 50
            cost_str = f"${p['shopify_cost']:.2f}" if p['shopify_cost'] is not None else "N/A"
            r.append(f"| {p['product_title'][:40]} | {p['sku']} | {cost_str} | {p['tags'][:30]} |")
        if len(zero_price) > 50:
            r.append(f"| ... and {len(zero_price) - 50} more |")
        r.append("")
        r.append("---")
        r.append("")

    # NO COST SET
    if no_cost:
        # Exclude zero-price items since they're likely drafts
        no_cost_with_price = [x for x in no_cost if x['retail_price'] > 0]
        r.append("## Missing Cost Field in Shopify")
        r.append("")
        r.append(f"**{len(no_cost_with_price)} variants** (with retail price > $0) have no cost set in Shopify.")
        r.append("Profitability cannot be tracked for these items.")
        r.append("")
        if no_cost_with_price:
            r.append("| Product | SKU | Retail Price | WYN Cost | Expected Cost |")
            r.append("|---------|-----|-------------|----------|---------------|")
            for p in no_cost_with_price[:50]:
                wyn_str = f"${p['wyn_landed_cost']:.2f}" if p['wyn_landed_cost'] else "N/A"
                exp_str = f"${p['expected_shopify_cost']:.2f}" if p['expected_shopify_cost'] else "N/A"
                r.append(f"| {p['product_title'][:40]} | {p['sku']} | ${p['retail_price']:.2f} | {wyn_str} | {exp_str} |")
            if len(no_cost_with_price) > 50:
                r.append(f"| ... and {len(no_cost_with_price) - 50} more |")
        r.append("")
        r.append("---")
        r.append("")

    # COST MISMATCH
    if cost_mismatch:
        r.append("## Cost Mismatches (Shopify vs Supplier Spreadsheet)")
        r.append("")
        r.append(f"**{len(cost_mismatch)} variants** have a Shopify cost that differs from the expected cost by > $1.00.")
        r.append("")
        r.append("| Product | SKU | Shopify Cost | WYN Cost | Expected | Diff |")
        r.append("|---------|-----|-------------|----------|----------|------|")
        for p in cost_mismatch[:50]:
            diff = p['shopify_cost'] - p['expected_shopify_cost']
            sign = '+' if diff > 0 else ''
            r.append(f"| {p['product_title'][:40]} | {p['sku']} | ${p['shopify_cost']:.2f} | ${p['wyn_landed_cost']:.2f} | ${p['expected_shopify_cost']:.2f} | {sign}${diff:.2f} |")
        r.append("")
        r.append("---")
        r.append("")

    # HEALTHY
    r.append("## Healthy Products (Margin >= 30%)")
    r.append("")
    r.append(f"{len(healthy)} variants meet or exceed the 30% margin target.")
    r.append("")

    # Margin distribution
    r.append("### Margin Distribution")
    r.append("")
    brackets = [
        ("Negative (< 0%)", [x for x in has_margin if x['margin_shopify'] < 0]),
        ("Critical (0-15%)", [x for x in has_margin if 0 <= x['margin_shopify'] < 15]),
        ("Low (15-30%)", [x for x in has_margin if 15 <= x['margin_shopify'] < 30]),
        ("Target (30-40%)", [x for x in has_margin if 30 <= x['margin_shopify'] < 40]),
        ("Good (40-50%)", [x for x in has_margin if 40 <= x['margin_shopify'] < 50]),
        ("Strong (50-60%)", [x for x in has_margin if 50 <= x['margin_shopify'] < 60]),
        ("Excellent (60%+)", [x for x in has_margin if x['margin_shopify'] >= 60]),
    ]
    r.append("| Margin Range | Count | % of Total |")
    r.append("|-------------|-------|-----------|")
    for label, items in brackets:
        pct = (len(items) / len(has_margin) * 100) if has_margin else 0
        r.append(f"| {label} | {len(items)} | {pct:.1f}% |")
    r.append("")

    r.append("---")
    r.append("")

    # Next steps
    r.append("## Action Items")
    r.append("")
    r.append("### Immediate")
    r.append("")
    if negative:
        r.append(f"1. **Fix {len(negative)} negative-margin products** — these are losing money on every sale")
    if low_margin:
        r.append(f"2. **Review {len(low_margin)} low-margin products** — raise prices or verify costs")
    if zero_price:
        r.append(f"3. **Check {len(zero_price)} $0 price variants** — these may be selling for free or are drafts")
    r.append("")
    r.append("### Short-Term")
    r.append("")
    if no_cost:
        no_cost_priced = len([x for x in no_cost if x['retail_price'] > 0])
        r.append(f"4. **Set cost for {no_cost_priced} variants** missing costs (run `npm run costs:execute`)")
    if cost_mismatch:
        r.append(f"5. **Resolve {len(cost_mismatch)} cost mismatches** — Shopify cost differs from supplier pricing")
    r.append("")
    r.append("---")
    r.append(f"*Report generated from live Shopify API data on {today}*")
    r.append(f"*Supplier reference: yhs_supply_products.xlsx ({len(supplier)} products)*")

    # Write report
    report_text = '\n'.join(r)
    report_path = f"MARGIN_AUDIT_REPORT_{today}.md"
    with open(report_path, 'w') as f:
        f.write(report_text)
    print(f"\n>> Report saved to: {report_path}")

    # Write JSON
    json_path = f"margin_audit_data_{today}.json"
    with open(json_path, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    print(f">> Raw data saved to: {json_path}")


if __name__ == '__main__':
    main()
